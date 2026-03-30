import json
import time
import urllib.request
import urllib.error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_FILE = "/mnt/d/0069085-260226173443078.csv"
OUTPUT_FILE = "japan_fungi_top300.xlsx"
TOP_N       = 300    # ← ここを変更するだけで種数を調整できます

# ---- 有名種ボーナスリスト（出現件数に加算するスコア）-------------------
# 食用・毒・商業栽培種など人間との関わりが深い種を優先するためのリスト
PRIORITY_BONUS = {
    # 食用（よく知られた種）
    "Lentinula edodes":        10000,  # シイタケ
    "Grifola frondosa":        10000,  # マイタケ
    "Flammulina velutipes":    10000,  # エノキタケ
    "Pleurotus ostreatus":     10000,  # ヒラタケ
    "Hypsizygus marmoreus":    10000,  # ブナシメジ
    "Pholiota nameko":         10000,  # ナメコ
    "Tricholoma matsutake":    10000,  # マツタケ
    "Agaricus bisporus":       10000,  # マッシュルーム
    "Auricularia auricula-judae": 9000,# キクラゲ
    "Tremella fuciformis":      9000,  # シロキクラゲ
    "Sparassis crispa":         9000,  # ハナビラタケ
    "Hericium erinaceus":       9000,  # ヤマブシタケ
    "Lyophyllum shimeji":       9000,  # ホンシメジ
    "Boletus edulis":           8000,  # ヤマドリタケ
    "Cantharellus cibarius":    8000,  # アンズタケ
    "Armillaria mellea":        8000,  # ナラタケ
    "Morchella esculenta":      8000,  # アミガサタケ
    "Lactarius volemus":        7000,  # チチタケ
    "Ganoderma lucidum":        7000,  # マンネンタケ（霊芝）
    "Suillus luteus":           7000,  # ヌメリイグチ
    # 毒キノコ（認知度が高い）
    "Amanita muscaria":        10000,  # ベニテングタケ
    "Amanita phalloides":      10000,  # タマゴテングタケ
    "Amanita virosa":          10000,  # ドクツルタケ
    "Amanita pantherina":       9000,  # ヒョウモンテングタケ
    "Omphalotus japonicus":     9000,  # ツキヨタケ
    "Hypholoma fasciculare":    9000,  # ニガクリタケ
    "Entoloma rhodopolium":     8000,  # イッポンシメジ
    "Rubroboletus satanas":     8000,  # ドクヤマドリ
    "Gyromitra esculenta":      8000,  # シャグマアミガサタケ
    "Cortinarius orellanus":    7000,  # チャイロフウセンタケ
}

# ---- 分類階層の日本語マッピング（前回と同じ）---------------------------
ORDER_JA = {
    "Agaricales": "ハラタケ目", "Boletales": "イグチ目",
    "Russulales": "ベニタケ目", "Polyporales": "タコウキン目",
    "Cantharellales": "アンズタケ目", "Auriculariales": "キクラゲ目",
    "Tremellales": "シロキクラゲ目", "Dacrymycetales": "アカキクラゲ目",
    "Pezizales": "チャワンタケ目", "Hypocreales": "ニクザキン目",
    "Phallales": "スッポンタケ目", "Geastrales": "ツチグリ目",
    "Hymenochaetales": "サビアナタケ目", "Thelephorales": "テレフォラレス目",
    "Gomphales": "ホウキタケ目", "Xylariales": "クロサイワイタケ目",
}
FAMILY_JA = {
    "Amanitaceae": "テングタケ科", "Agaricaceae": "ハラタケ科",
    "Tricholomataceae": "キシメジ科", "Lyophyllaceae": "シメジ科",
    "Marasmiaceae": "オキナタケ科", "Physalacriaceae": "モエギタケ科",
    "Pleurotaceae": "ヒラタケ科", "Strophariaceae": "モエギタケ科",
    "Entolomataceae": "イッポンシメジ科", "Cortinariaceae": "フウセンタケ科",
    "Inocybaceae": "アセタケ科", "Psathyrellaceae": "ナヨタケ科",
    "Omphalotaceae": "ツキヨタケ科", "Hygrophoraceae": "ヌメリガサ科",
    "Russulaceae": "ベニタケ科", "Boletaceae": "イグチ科",
    "Suillaceae": "ヌメリイグチ科", "Polyporaceae": "タコウキン科",
    "Meripilaceae": "マイタケ科", "Fomitopsidaceae": "サルノコシカケ科",
    "Sparassidaceae": "シロアミタケ科", "Cantharellaceae": "アンズタケ科",
    "Hydnaceae": "ヒダナシタケ科", "Clavulinaceae": "エセオリミキ科",
    "Clavariaceae": "シロソウメンタケ科", "Auriculariaceae": "キクラゲ科",
    "Tremellaceae": "シロキクラゲ科", "Dacrymycetaceae": "アカキクラゲ科",
    "Morchellaceae": "アミガサタケ科", "Helvellaceae": "シャグマアミガサタケ科",
    "Pezizaceae": "チャワンタケ科", "Sarcoscyphaceae": "サカズキキン科",
    "Phallaceae": "スッポンタケ科", "Geastraceae": "ツチグリ科",
    "Hymenochaetaceae": "サビアナタケ科", "Ramariaceae": "ホウキタケ科",
    "Hericiaceae": "サンゴハリタケ科", "Bankeraceae": "ニセショウロ科",
    "Bondarzewiaceae": "ボンダルツェウィア科",
}
GENUS_JA = {
    "Amanita": "テングタケ属", "Tricholoma": "キシメジ属",
    "Lyophyllum": "シメジ属", "Hypsizygus": "ブナシメジ属",
    "Lentinula": "シイタケ属", "Flammulina": "エノキタケ属",
    "Armillaria": "ナラタケ属", "Pleurotus": "ヒラタケ属",
    "Pholiota": "ナメコ属", "Hypholoma": "クリタケ属",
    "Stropharia": "モエギタケ属", "Entoloma": "イッポンシメジ属",
    "Cortinarius": "フウセンタケ属", "Hebeloma": "ワカフサタケ属",
    "Inocybe": "アセタケ属", "Agaricus": "ハラタケ属",
    "Lepiota": "カラカサタケ属", "Macrolepiota": "オニカラカサタケ属",
    "Calvatia": "オニフスベ属", "Lycoperdon": "ホコリタケ属",
    "Coprinus": "ヒトヨタケ属", "Coprinellus": "キラタケ属",
    "Coprinopsis": "ヒトヨタケ属", "Psathyrella": "ナヨタケ属",
    "Omphalotus": "ツキヨタケ属", "Gymnopus": "クヌギタケ属",
    "Marasmius": "クヌギタケ属", "Russula": "ベニタケ属",
    "Lactarius": "チチタケ属", "Lactifluus": "チチタケ属",
    "Boletus": "ヤマドリタケ属", "Neoboletus": "アカヤマドリ属",
    "Rubroboletus": "ドクヤマドリ属", "Leccinum": "ヤマドリタケモドキ属",
    "Suillus": "ヌメリイグチ属", "Tylopilus": "ニガイグチ属",
    "Ganoderma": "マンネンタケ属", "Trametes": "カワラタケ属",
    "Polyporus": "チチタケ属", "Grifola": "マイタケ属",
    "Laetiporus": "ニワトリタケ属", "Sparassis": "ハナビラタケ属",
    "Cantharellus": "アンズタケ属", "Craterellus": "クロラッパタケ属",
    "Clavulina": "エセオリミキ属", "Ramaria": "ホウキタケ属",
    "Hericium": "サンゴハリタケ属", "Hydnum": "ヒダナシタケ属",
    "Auricularia": "キクラゲ属", "Tremella": "シロキクラゲ属",
    "Morchella": "アミガサタケ属", "Helvella": "シャグマアミガサタケ属",
    "Phallus": "スッポンタケ属", "Dictyophora": "キヌガサタケ属",
    "Geastrum": "ツチグリ属",
}
PHYLUM_JA = {
    "Basidiomycota": "担子菌門", "Ascomycota": "子嚢菌門",
}
CLASS_JA = {
    "Agaricomycetes": "ハラタケ綱", "Dacrymycetes": "アカキクラゲ綱",
    "Tremellomycetes": "シロキクラゲ綱", "Pezizomycetes": "チャワンタケ綱",
    "Sordariomycetes": "フンタマカビ綱", "Leotiomycetes": "レオチオミセス綱",
}
KINGDOM_JA = {"Fungi": "菌界", "Plantae": "植物界", "Animalia": "動物界"}

# ---- GBIF API 和名取得 -------------------------------------------------
def fetch_japanese_name(taxon_key: str) -> str:
    if not taxon_key:
        return ""
    url = f"https://api.gbif.org/v1/species/{taxon_key}/vernacularNames?limit=20"
    try:
        with urllib.request.urlopen(url, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for item in data.get("results", []):
            if item.get("language", "").lower() in ("ja", "jpn", "japanese"):
                return item.get("vernacularName", "")
    except Exception:
        pass
    return ""

# ---- TSV パーサー -------------------------------------------------------
def parse_tsv_line(line: str) -> list[str]:
    return line.rstrip("\n").split("\t")

# ========================================================================
# パス1：全行を走査してスコアを集計
# ========================================================================
print("パス1: スコア集計中...")

# species -> {"score": int, "row": dict}
species_data: dict[str, dict] = {}

with open(INPUT_FILE, encoding="utf-8") as f:
    col_index = {name: i for i, name in enumerate(parse_tsv_line(f.readline()))}

    def get(cols, name):
        i = col_index.get(name)
        return cols[i].strip() if i is not None and i < len(cols) else ""

    for line in f:
        cols = parse_tsv_line(line)

        if get(cols, "kingdom") != "Fungi":
            continue
        if get(cols, "taxonRank") != "SPECIES":
            continue
        status = get(cols, "taxonomicStatus")
        if status not in ("ACCEPTED", ""):
            continue

        species = get(cols, "species")
        if not species:
            continue

        # 出現件数を1件としてカウント（行 = 1 occurrence record）
        if species not in species_data:
            species_data[species] = {
                "score": 0,
                "kingdom":  get(cols, "kingdom"),
                "phylum":   get(cols, "phylum"),
                "class":    get(cols, "class"),
                "order":    get(cols, "order"),
                "family":   get(cols, "family"),
                "genus":    get(cols, "genus"),
                "sci":      get(cols, "scientificName"),
                "tkey":     get(cols, "taxonKey") or get(cols, "speciesKey"),
                "status":   status,
            }
        species_data[species]["score"] += 1

    # 有名種ボーナスを加算
    for sp, bonus in PRIORITY_BONUS.items():
        if sp in species_data:
            species_data[sp]["score"] += bonus

# 上位 TOP_N 種を選択
top_species = sorted(
    species_data.items(),
    key=lambda x: x[1]["score"],
    reverse=True
)[:TOP_N]

print(f"パス1完了: {len(species_data):,} 種中 上位 {len(top_species)} 種を選択")

# ========================================================================
# パス2：Excel 書き出し（+ API 和名取得）
# ========================================================================
print("パス2: Excel書き出し中...")

wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet("日本産キノコ分類")

OUT_HEADERS = [
    "rank",         # 出現頻度順位
    "score",        # スコア（出現件数＋ボーナス）
    "kingdom",      "kingdom_ja",
    "phylum",       "phylum_ja",
    "class",        "class_ja",
    "order",        "order_ja",
    "family",       "family_ja",
    "genus",        "genus_ja",
    "species",
    "scientific_name",
    "japanese_name",
    "taxonomic_status",
    "taxon_key",
]

header_fill = PatternFill("solid", start_color="2E4057")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
data_font   = Font(name="Arial", size=9)
alt_fill    = PatternFill("solid", start_color="F5F7FA")

def header_cell(v):
    c = openpyxl.cell.WriteOnlyCell(ws, value=v)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def data_cell(v, idx):
    c = openpyxl.cell.WriteOnlyCell(ws, value=v)
    c.font = data_font
    if idx % 2 == 0:
        c.fill = alt_fill
    return c

ws.append([header_cell(h) for h in OUT_HEADERS])

for rank, (species, d) in enumerate(top_species, 1):
    ja_name = fetch_japanese_name(d["tkey"])
    if d["tkey"]:
        time.sleep(0.05)

    row = [
        rank,
        d["score"],
        d["kingdom"],  KINGDOM_JA.get(d["kingdom"], ""), # d["kingdom"],  PHYLUM_JA.get(d["kingdom"], ""),
        d["phylum"],   PHYLUM_JA.get(d["phylum"], ""),
        d["class"],    CLASS_JA.get(d["class"], ""),
        d["order"],    ORDER_JA.get(d["order"], ""),
        d["family"],   FAMILY_JA.get(d["family"], ""),
        d["genus"],    GENUS_JA.get(d["genus"], ""),
        species,
        d["sci"],
        ja_name,
        d["status"],
        d["tkey"],
    ]
    ws.append([data_cell(v, rank) for v in row])

    if rank % 50 == 0:
        print(f"  書き出し済み: {rank} / {TOP_N}")

wb.save(OUTPUT_FILE)
print(f"\n完了 → {OUTPUT_FILE}")