# GBIFからダウンロードしたCSVを読み込み
INPUT_FILE = "/mnt/d/0069085-260226173443078.csv"

import json
import time
import urllib.request
import urllib.error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "japan_fungi_full.xlsx"

# ---- 分類階層の日本語マッピング ----------------------------------------
# 目 (order)
ORDER_JA = {
    "Agaricales": "ハラタケ目", "Boletales": "イグチ目",
    "Russulales": "ベニタケ目", "Polyporales": "タコウキン目",
    "Cantharellales": "アンズタケ目", "Auriculariales": "キクラゲ目",
    "Tremellales": "シロキクラゲ目", "Dacrymycetales": "アカキクラゲ目",
    "Pezizales": "チャワンタケ目", "Hypocreales": "ニクザキン目",
    "Xylariales": "クロサイワイタケ目", "Phallales": "スッポンタケ目",
    "Geastrales": "ツチグリ目", "Gomphales": "ホウキタケ目",
    "Hymenochaetales": "サビアナタケ目", "Thelephorales": "テレフォラレス目",
    "Sebacinales": "セバシナレス目", "Corticiales": "コルチシアレス目",
    "Gloeophyllales": "グロエオフィラレス目", "Trechisporales": "トレキスポラレス目",
}

# 科 (family)
FAMILY_JA = {
    "Amanitaceae": "テングタケ科", "Agaricaceae": "ハラタケ科",
    "Tricholomataceae": "キシメジ科", "Lyophyllaceae": "シメジ科",
    "Marasmiaceae": "オキナタケ科", "Physalacriaceae": "モエギタケ科",
    "Pleurotaceae": "ヒラタケ科", "Strophariaceae": "モエギタケ科",
    "Entolomataceae": "イッポンシメジ科", "Cortinariaceae": "フウセンタケ科",
    "Inocybaceae": "アセタケ科", "Psathyrellaceae": "ナヨタケ科",
    "Omphalotaceae": "ツキヨタケ科", "Hygrophoraceae": "ヌメリガサ科",
    "Crepidotaceae": "ウロコタケ科", "Russulaceae": "ベニタケ科",
    "Boletaceae": "イグチ科", "Suillaceae": "ヌメリイグチ科",
    "Paxillaceae": "キシメジ科", "Gomphidiaceae": "ツルタケ科",
    "Polyporaceae": "タコウキン科", "Meripilaceae": "マイタケ科",
    "Fomitopsidaceae": "サルノコシカケ科", "Sparassidaceae": "シロアミタケ科",
    "Bondarzewiaceae": "ボンダルツェウィア科",
    "Cantharellaceae": "アンズタケ科", "Hydnaceae": "ヒダナシタケ科",
    "Clavulinaceae": "エセオリミキ科", "Clavariaceae": "シロソウメンタケ科",
    "Auriculariaceae": "キクラゲ科", "Tremellaceae": "シロキクラゲ科",
    "Dacrymycetaceae": "アカキクラゲ科", "Morchellaceae": "アミガサタケ科",
    "Helvellaceae": "シャグマアミガサタケ科", "Pezizaceae": "チャワンタケ科",
    "Sarcoscyphaceae": "サカズキキン科", "Phallaceae": "スッポンタケ科",
    "Geastraceae": "ツチグリ科", "Hymenochaetaceae": "サビアナタケ科",
    "Ramariaceae": "ホウキタケ科", "Hericiaceae": "サンゴハリタケ科",
    "Bankeraceae": "ニセショウロ科", "Clavulinaceae": "エセオリミキ科",
    "Gloeophyllaceae": "チャウロコタケ科",
}

# 属 (genus) — 主要属のみ。未登録はそのまま英語表記
GENUS_JA = {
    "Amanita": "テングタケ属", "Tricholoma": "キシメジ属",
    "Lyophyllum": "シメジ属", "Hypsizygus": "ブナシメジ属",
    "Lentinula": "シイタケ属", "Flammulina": "エノキタケ属",
    "Armillaria": "ナラタケ属", "Pleurotus": "ヒラタケ属",
    "Pholiota": "ナメコ属", "Hypholoma": "クリタケ属",
    "Stropharia": "モエギタケ属", "Entoloma": "イッポンシメジ属",
    "Cortinarius": "フウセンタケ属", "Hebeloma": "ワカフサタケ属",
    "Inocybe": "アセタケ属", "Agaricus": "ハラタケ属",
    "Lepiota": "カラカサタケ属", "Macrolepiota": "オニカラカサタケ属",
    "Calvatia": "オニフスベ属", "Lycoperdon": "ホコリタケ属",
    "Coprinus": "ヒトヨタケ属", "Coprinellus": "キラタケ属",
    "Coprinopsis": "ヒトヨタケ属", "Psathyrella": "ナヨタケ属",
    "Omphalotus": "ツキヨタケ属", "Gymnopus": "クヌギタケ属",
    "Marasmius": "クヌギタケ属", "Russula": "ベニタケ属",
    "Lactarius": "チチタケ属", "Lactifluus": "チチタケ属",
    "Boletus": "ヤマドリタケ属", "Neoboletus": "アカヤマドリ属",
    "Rubroboletus": "ドクヤマドリ属", "Leccinum": "ヤマドリタケモドキ属",
    "Suillus": "ヌメリイグチ属", "Xerocomus": "キクバナイグチ属",
    "Tylopilus": "ニガイグチ属", "Gyroporus": "マイタケイグチ属",
    "Strobilomyces": "ガンタケイグチ属", "Ganoderma": "マンネンタケ属",
    "Trametes": "カワラタケ属", "Fomes": "ツリガネタケ属",
    "Polyporus": "チチタケ属", "Grifola": "マイタケ属",
    "Laetiporus": "ニワトリタケ属", "Fomitopsis": "ベニサルノコシカケ属",
    "Sparassis": "ハナビラタケ属", "Inonotus": "ブナサルノコシカケ属",
    "Phellinus": "ツガサルノコシカケ属", "Cantharellus": "アンズタケ属",
    "Craterellus": "クロラッパタケ属", "Clavulina": "エセオリミキ属",
    "Clavaria": "シロソウメンタケ属", "Ramaria": "ホウキタケ属",
    "Hericium": "サンゴハリタケ属", "Hydnum": "ヒダナシタケ属",
    "Hydnellum": "クロカワ属", "Sarcodon": "ニセショウロ属",
    "Auricularia": "キクラゲ属", "Tremella": "シロキクラゲ属",
    "Calocera": "ツノマタタケ属", "Dacryopinax": "スエヒロタケ属",
    "Morchella": "アミガサタケ属", "Helvella": "シャグマアミガサタケ属",
    "Gyromitra": "シャグマアミガサタケ属", "Peziza": "チャワンタケ属",
    "Sarcoscypha": "サカズキキン属", "Phallus": "スッポンタケ属",
    "Dictyophora": "キヌガサタケ属", "Mutinus": "コタケスッポンタケ属",
    "Geastrum": "ツチグリ属", "Bondarzewia": "ボンダルツェウィア属",
}

# 門 (phylum)
PHYLUM_JA = {
    "Basidiomycota": "担子菌門", "Ascomycota": "子嚢菌門",
    "Chytridiomycota": "ツボカビ門", "Zygomycota": "接合菌門",
    "Mucoromycota": "ケカビ門",
}

# 綱 (class)
CLASS_JA = {
    "Agaricomycetes": "ハラタケ綱", "Dacrymycetes": "アカキクラゲ綱",
    "Tremellomycetes": "シロキクラゲ綱", "Auriculariales": "キクラゲ綱",
    "Pezizomycetes": "チャワンタケ綱", "Sordariomycetes": "フンタマカビ綱",
    "Dothideomycetes": "ドチデオミセス綱", "Leotiomycetes": "レオチオミセス綱",
    "Eurotiomycetes": "ユーロチオミセス綱",
}

# ---- GBIF Species API で和名を取得 (urllib のみ使用) --------------------
def fetch_japanese_name(taxon_key: str) -> str:
    """Return Japanese vernacular name from GBIF Species API, or empty string."""
    if not taxon_key:
        return ""
    url = f"https://api.gbif.org/v1/species/{taxon_key}/vernacularNames?limit=20"
    try:
        with urllib.request.urlopen(url, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for item in data.get("results", []):
            if item.get("language", "").lower() in ("ja", "jpn", "japanese"):
                return item.get("vernacularName", "")
    except (urllib.error.URLError, json.JSONDecodeError, KeyError):
        pass
    return ""

# ---- TSV パーサー -------------------------------------------------------
def parse_tsv_line(line: str) -> list[str]:
    return line.rstrip("\n").split("\t")

# ---- Excel セル ファクトリ ---------------------------------------------
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet("日本産キノコ分類")

OUT_HEADERS_EN = [
    "kingdom",    "kingdom_ja",
    "phylum",     "phylum_ja",
    "class",      "class_ja",
    "order",      "order_ja",
    "family",     "family_ja",
    "genus",      "genus_ja",
    "species",    "species_ja",
    "scientific_name", "japanese_name",
    "taxon_rank", "taxonomic_status", "taxon_key",
]

header_fill = PatternFill("solid", start_color="2E4057")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
data_font   = Font(name="Arial", size=9)
alt_fill    = PatternFill("solid", start_color="F5F7FA")

def header_cell(value):
    c = openpyxl.cell.WriteOnlyCell(ws, value=value)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def data_cell(value, row_idx):
    c = openpyxl.cell.WriteOnlyCell(ws, value=value)
    c.font = data_font
    if row_idx % 2 == 0:
        c.fill = alt_fill
    return c

ws.append([header_cell(h) for h in OUT_HEADERS_EN])

# ---- メイン処理 ---------------------------------------------------------
seen_species = set()
row_idx = 0
api_errors = 0

with open(INPUT_FILE, encoding="utf-8") as f:
    col_index = {name: i for i, name in enumerate(parse_tsv_line(f.readline()))}

    def get(cols, name):
        i = col_index.get(name)
        return cols[i].strip() if i is not None and i < len(cols) else ""

    for line in f:
        cols = parse_tsv_line(line)

        if get(cols, "taxonRank") != "SPECIES":
            continue
        status = get(cols, "taxonomicStatus")
        if status not in ("ACCEPTED", ""):
            continue

        species = get(cols, "species")
        if not species or species in seen_species:
            continue
        seen_species.add(species)

        kingdom = get(cols, "kingdom")
        phylum  = get(cols, "phylum")
        cls     = get(cols, "class")
        order   = get(cols, "order")
        family  = get(cols, "family")
        genus   = get(cols, "genus")
        sci     = get(cols, "scientificName")
        tkey    = get(cols, "taxonKey") or get(cols, "speciesKey")

        # 日本語和名をAPIから取得（rate limit対策で少し待つ）
        ja_name = fetch_japanese_name(tkey)
        if tkey:
            time.sleep(0.05)   # 20 req/sec 以内に抑える

        values = [
            kingdom, PHYLUM_JA.get(kingdom, ""),   # kingdom / kingdom_ja
            phylum,  PHYLUM_JA.get(phylum, ""),    # phylum  / phylum_ja
            cls,     CLASS_JA.get(cls, ""),         # class   / class_ja
            order,   ORDER_JA.get(order, ""),       # order   / order_ja
            family,  FAMILY_JA.get(family, ""),     # family  / family_ja
            genus,   GENUS_JA.get(genus, ""),       # genus   / genus_ja
            species, "",                            # species / species_ja (英語のみ)
            sci,     ja_name,                       # scientific_name / japanese_name
            get(cols, "taxonRank"),
            status,
            tkey,
        ]

        ws.append([data_cell(v, row_idx) for v in values])
        row_idx += 1

        if row_idx % 500 == 0:
            print(f"  処理済み: {row_idx:,} 種  (API取得済み和名あり)")

wb.save(OUTPUT_FILE)
print(f"\n完了: {row_idx:,} 種 → {OUTPUT_FILE}")